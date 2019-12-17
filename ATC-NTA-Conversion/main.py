import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

def processextract(perioddata, start, dirs, classes):
        
        # create new dataframe for period results
        dfperiod = pd.DataFrame(columns=hdrs)
        
        # get unclassified 85th percentile speed for period
        perc85DirA = np.percentile(perioddata['Speed'][perioddata['Dr'] == dirs[0]], 85)
        perc85DirB = np.percentile(perioddata['Speed'][perioddata['Dr'] == dirs[1]], 85)
        
        # loop through classes and populate count, ave speed, percentile for each vehicle type
        for veh in classes:
                dfveh = pd.DataFrame(columns=hdrs)
                vehdata = perioddata[perioddata['Class'] == veh]                
                
                # add site date, start/end time and vehicle type
                dfveh['Date']= perioddata['YYYY-MM-DD'].unique()
                dfveh['Start Time'] = start.time()
                dfveh['End Time'] = (start+timedelta(minutes=15)).time()
                dfveh['Vehicle Type'] = veh
                        
                # extract data for each direction
                dfdirA = vehdata[vehdata['Dr'] ==  dirs[0]]
                dfdirB = vehdata[vehdata['Dr'] == dirs[1]]
                
                dfveh['DIRA'] = len(dfdirA)
                dfveh['DIRB'] = len(dfdirB)
                
                # get average speed for vehicle for each direction
                dfveh['DIRA AVG SPEED']= dfdirA['Speed'].mean()
                dfveh['DIRB AVG SPEED']= dfdirB['Speed'].mean()
        
                if len(dfdirA)>0:
                        dfveh['DIRA 85TH PERCENTILE / UC'] = perc85DirA
                
                if len(dfdirB)>0:
                        dfveh['DIRB 85TH PERCENTILE / UC'] = perc85DirB
                        
                dfperiod = dfperiod.append(dfveh)
                
        return dfperiod

def daterange(date1, date2):
    for n in range(int ((date2 - date1).days)+1):
        yield date1 + timedelta(n)

#  raw data headers - YYYY-MM-DD        hh:mm:ss        Dr      Speed   Cl

# vehicle classes
allclasses = ['M/C','Car','LGV','PSV','OGV1','OGV1','OGV2','OGV2','OGV2','OGV2','OGV2','OGV2','OGV2']
classes = ['CAR','LGV','OGV1','OGV2','PSV','M/C']

# headers for results
hdrs = "Site No,Index,Date,Start Time,End Time,Vehicle Type,DIRA,DIRB,DIRA AVG SPEED,DIRB AVG SPEED,DIRA 85TH PERCENTILE / UC,DIRB 85TH PERCENTILE / UC".split(',')

filename = 'SITE1_ALLRAWDATA.xlsx'

print(str(datetime.now()))

dfdict={}

# create data frame for results
dfresults = pd.DataFrame(columns=hdrs)

# read raw data
xls_file = pd.ExcelFile(filename)

raw_data = xls_file.parse('RawData')

raw_data['Class'] = raw_data['Cl'].apply(lambda x: allclasses[x-1].upper())

raw_data['Time'] = raw_data.apply(lambda r : pd.datetime.combine(r['YYYY-MM-DD'],r['hh:mm:ss']),1)

#print(raw_data)

#raw_input("Press the <ENTER> key to continue...")

# get start/end dates of survey from raw data
start_date=raw_data['YYYY-MM-DD'].min()

end_date=raw_data['YYYY-MM-DD'].max()

# calculate no of days survey covered
days = (end_date-start_date).days+1 #  calculate from start/end

# get the directions from the raw data
dirs = raw_data['Dr'].unique()
#print(dirs)

#print('From {0} to {1}'.format(start_date, end_date))
#print (date_rng)

# split data out by date and time period
for dt in daterange(start_date, end_date):
        
  period_rng = pd.date_range(start=dt, periods=96, freq='15min')
  
  surveydate = dt
  
  dateextract = raw_data[(raw_data['YYYY-MM-DD'] == surveydate)]
  
  if dateextract.empty:
    dateextract = raw_data[(raw_data['YYYY-MM-DD'] == surveydate-timedelta(1))].copy()
    dateextract['YYYY-MM-DD'] = surveydate
    dateextract['Speed'] = 0
    dateextract['Class'] = ''
    dateextract['Time'] = dateextract['Time'].apply(lambda t : t + timedelta(1))
  
  #dfdict[str(surveydate)] = dateextract
  
  #print(list(dfdict))
  
  #raw_input("Press the <ENTER> key to continue...")
  
  if len(dateextract)>0:
    
    print(dt)
    
    for tm in period_rng:
      start= tm
      finish =tm+timedelta(minutes=15)
      #print("{0}-{1}".format(start, finish))

      #raw_input("Press the <ENTER> key to continue...                  

      periodextract = dateextract[(dateextract['Time']>=start) & (dateextract['Time']<finish)]
      
      if len(periodextract)>0:
        #print(periodextract)

        #raw_input("Press the <ENTER> key to continue...")

        periodresults = processextract(periodextract, tm, dirs, classes)

        #print(periodresults)

        dfresults = dfresults.append(periodresults)
        
        #raw_input("Press the <ENTER> key to continue...")

dfresults = dfresults.fillna('-')

dfresults['Site No'] = 'Test'   

dfresults['Index'] = range(1, len(dfresults)+1)

#print(dfresults.to_string())

# load destination workbook
wb = load_workbook(filename)

# add new sheet for results
ws_results = wb.create_sheet(title='Data')

# output results to new sheet
for r in dataframe_to_rows(dfresults, index=False, header=True):
  ws_results.append(r)

# format sheet
  
# save 
wb.save(filename)

with pd.ExcelWriter(r'SITE1_ALLRAWDATA-results.xlsx') as writer:
    dfresults.to_excel(writer, sheet_name='Data', float_format='%.2f', index=False)

print(str(datetime.now()))
